import pandas as pd
import openpyxl
import xlrd

# 1. Try loading with different methods to check indexing
print("CHECKING EXCEL FILE USING MULTIPLE METHODS:\n")

# Method 1: Using pandas
print("Method 1: Using pandas")
try:
    df = pd.read_excel('MASTER COPY.xlsx', sheet_name='MASTER COPY')
    print(f"Pandas detected {len(df)} rows (0-based index)")
    print(f"Last row index: {len(df)-1}")
    print(f"After converting to 1-based indexing, last row would be: {len(df)}")
    
    # Try to manually check some of the rows that should exist
    rows_to_check = [14790, 14811, 14812]
    
    # First check with pandas default 0-based indexing
    print("\nChecking with pandas 0-based indexing:")
    for row in rows_to_check:
        zero_idx = row - 1  # Convert to 0-based
        if zero_idx < len(df) and zero_idx >= 0:
            desc = df.iloc[zero_idx].get('description', 'N/A')
            print(f"Row {row} (0-based index {zero_idx}): Description = {desc}")
        else:
            print(f"Row {row} (0-based index {zero_idx}): Out of range")
    
    # Now convert to 1-based and try loc[]
    print("\nConverting to 1-based indexing and checking:")
    df.index = df.index + 1
    for row in rows_to_check:
        if row in df.index:
            desc = df.loc[row].get('description', 'N/A')
            print(f"Row {row}: Description = {desc}")
        else:
            print(f"Row {row}: Not found in index")
    
except Exception as e:
    print(f"Error with pandas: {e}")

# Method 2: Using openpyxl to directly access worksheet cells
print("\nMethod 2: Using openpyxl")
try:
    wb = openpyxl.load_workbook('MASTER COPY.xlsx', read_only=True)
    sheet = wb['MASTER COPY']
    print(f"openpyxl detected {sheet.max_row} rows")
    
    # openpyxl is 1-based, so we can check directly
    for row in rows_to_check:
        if row <= sheet.max_row:
            # Find the column with description (assuming it's column A)
            # Try a few columns to find the description
            desc = None
            for col in range(1, 10):  # Try first 10 columns
                cell_value = sheet.cell(row=row, column=col).value
                if isinstance(cell_value, str) and len(cell_value) > 5:
                    desc = cell_value
                    break
            
            print(f"Row {row}: Found in sheet, possible description = {desc}")
        else:
            print(f"Row {row}: Beyond max row {sheet.max_row}")
    
    wb.close()
except Exception as e:
    print(f"Error with openpyxl: {e}")

# Method 3: Using xlrd for xls files
print("\nMethod 3: Using xlrd")
try:
    wb = xlrd.open_workbook('MASTER COPY.xlsx')
    sheet = wb.sheet_by_name('MASTER COPY')
    print(f"xlrd detected {sheet.nrows} rows (0-based index)")
    
    # xlrd is 0-based
    for row in rows_to_check:
        zero_idx = row - 1
        if zero_idx < sheet.nrows:
            # Try to find description in first few cells
            row_values = sheet.row_values(zero_idx)
            potential_desc = [val for val in row_values if isinstance(val, str) and len(val) > 5]
            desc = potential_desc[0] if potential_desc else "No string found"
            print(f"Row {row} (0-based index {zero_idx}): {desc}")
        else:
            print(f"Row {row} (0-based index {zero_idx}): Beyond max row {sheet.nrows-1}")
except Exception as e:
    print(f"Error with xlrd: {e}")

# Check if this is possibly a multi-sheet issue
print("\nChecking all sheets in the Excel file:")
try:
    xls = pd.ExcelFile('MASTER COPY.xlsx')
    sheets = xls.sheet_names
    print(f"Found sheets: {sheets}")
    
    for sheet in sheets:
        df = pd.read_excel('MASTER COPY.xlsx', sheet_name=sheet)
        print(f"Sheet '{sheet}' has {len(df)} rows")
        
        # Check if our target rows exist in this sheet
        if len(df) >= 14812:
            print(f"  Sheet '{sheet}' contains row 14812!")
            print(f"  Value at row 14812: {df.iloc[14811].values}")  # 0-based index
except Exception as e:
    print(f"Error checking sheets: {e}")

# Output the exact name of the sheet we're trying to access
try:
    xls = pd.ExcelFile('MASTER COPY.xlsx')
    sheet_names = xls.sheet_names
    print(f"\nExact sheet names in the file: {sheet_names}")
    
    # Check if 'MASTER COPY' is exactly in the list
    if 'MASTER COPY' in sheet_names:
        print("Sheet 'MASTER COPY' exists with exact name match")
    else:
        print("No exact match for 'MASTER COPY', trying case-insensitive match")
        for name in sheet_names:
            if name.upper() == 'MASTER COPY':
                print(f"Found case-insensitive match: '{name}'")
except Exception as e:
    print(f"Error checking sheet names: {e}") 