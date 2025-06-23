import pandas as pd
import openpyxl

# First, check the raw Excel file with openpyxl to confirm row 14812 exists
print("Checking original Excel file with openpyxl:")
wb = openpyxl.load_workbook('MASTER COPY.xlsx', read_only=True)
sheet = wb['MASTER COPY']
max_row = sheet.max_row
print(f"Max row in MASTER COPY: {max_row}")

# Get values from row 14812
row_14812_values = []
for col in range(1, 20):  # Check first 20 columns
    value = sheet.cell(row=14812, column=col).value
    if value is not None:
        row_14812_values.append(value)

print(f"Values in row 14812: {row_14812_values}")
wb.close()

# Now check our generated output to see if the data from row 14812 is included
print("\nChecking output file:")
output_df = pd.read_excel('test_openpyxl_fix.xlsx')
print(f"Total rows in output: {len(output_df)}")

# Extract unique sizes and finishes
unique_sizes = output_df['Option1 Value'].unique()
unique_finishes = output_df['Option2 Value'].unique()
unique_skus = output_df['Variant SKU'].unique()

print(f"Unique sizes: {len(unique_sizes)}")
print(f"Unique finishes: {len(unique_finishes)}")
print(f"Unique SKUs: {len(unique_skus)}")
print(f"Unique SKUs: {unique_skus}")

# Check for specific values from row 14812
print("\nChecking if row 14812 data is in output:")
if any('17581/3X' in str(sku) for sku in unique_skus):
    print("Found SKU from row 14812 (17581/3X) in the output!")
else:
    print("Did not find SKU from row 14812 in the output.")
    
# Look for any rows with X## finish code
print("\nLooking for X## finish data in output:")
for idx, row in output_df.iterrows():
    if '17581/3X' in str(row['Variant SKU']):
        print(f"Found row with SKU 17581/3X: {row['Option1 Value']} - {row['Option2 Value']} - Price: {row['Variant Price']}")
        break 