import openpyxl
import os
import glob

def fix_excel_file(input_file=None, output_file=None):
    """
    Directly edit the Excel file to fix boolean values
    """
    # If no input file is specified, use the most recent generated file
    if not input_file:
        output_files = glob.glob('shopify_feed_*.xlsx')
        if not output_files:
            print("No output files found")
            return
        input_file = max(output_files, key=os.path.getmtime)
    
    print(f"Processing file: {input_file}")
    
    # If no output file is specified, create one
    if not output_file:
        output_file = 'final_' + os.path.basename(input_file)
    
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
    
    # Get headers
    headers = [cell.value for cell in sheet[1]]
    
    # Define boolean column indexes
    boolean_columns = ['Published', 'Variant Requires Shipping', 'Variant Taxable', 'Gift Card']
    boolean_indexes = [headers.index(col) + 1 for col in boolean_columns if col in headers]
    
    # Define regional inclusion column indexes
    inclusion_columns = [col for col in headers if str(col).startswith('Included /')]
    inclusion_indexes = [headers.index(col) + 1 for col in inclusion_columns]
    
    # Combine all boolean-type column indexes
    all_boolean_indexes = boolean_indexes + inclusion_indexes
    
    # Process rows
    for row_idx in range(2, sheet.max_row + 1):
        for col_idx in all_boolean_indexes:
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value is True:
                cell.value = "TRUE"
            elif cell.value is False:
                cell.value = "FALSE"
    
    # Save the workbook
    wb.save(output_file)
    print(f"Saved final file to {output_file}")
    
    # Identify premium finishes and check price variation
    premium_keywords = ['Polished Nickel', 'Brushed Nickel', 'Antique', 'FFPN', 'FFSN', 'FFAB']
    
    # Get Option2 Value column index
    finish_col_idx = headers.index('Option2 Value') + 1
    price_col_idx = headers.index('Variant Price') + 1
    
    # Check some values
    print("\nSample of values in the Excel file:")
    
    # Check boolean values
    for col_name in boolean_columns:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            values = []
            for row_idx in range(2, min(5, sheet.max_row + 1)):
                values.append(sheet.cell(row=row_idx, column=col_idx).value)
            print(f"{col_name}: {values}")
    
    # Check price variation
    premium_prices = []
    regular_prices = []
    
    for row_idx in range(2, sheet.max_row + 1):
        finish = sheet.cell(row=row_idx, column=finish_col_idx).value
        price = sheet.cell(row=row_idx, column=price_col_idx).value
        
        is_premium = any(keyword in str(finish) for keyword in premium_keywords)
        if is_premium:
            premium_prices.append(price)
        else:
            regular_prices.append(price)
    
    print(f"\nPremium finishes average price: {sum(premium_prices)/len(premium_prices)}")
    print(f"Regular finishes average price: {sum(regular_prices)/len(regular_prices)}")
    print(f"Price difference: {(sum(premium_prices)/len(premium_prices)) - (sum(regular_prices)/len(regular_prices))}")
    
if __name__ == "__main__":
    fix_excel_file() 